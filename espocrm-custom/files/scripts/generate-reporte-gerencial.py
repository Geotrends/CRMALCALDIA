#!/usr/bin/env python3
"""Genera reporte gerencial en Excel o HTML (para PDF)."""

from __future__ import annotations

import argparse
import html
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    Workbook = None


def esc(value) -> str:
    return html.escape(str(value if value is not None else ""))


def logo_html(data: dict) -> str:
    b64 = data.get("logoBase64")
    mime = data.get("logoMime") or "image/png"

    if not b64:
        return ""

    return (
        f'<img src="data:{mime};base64,{b64}" '
        'alt="Alcaldía de Envigado" width="64" height="64" '
        'style="width:64px;height:64px;" />'
    )


def build_html(data: dict) -> str:
    kpis = data.get("kpis", {})
    filtro = data.get("filtroAsignado")
    filtro_row = (
        f"<tr><td colspan='2' class='meta-cell'>"
        f"<span class='meta-label'>Filtro</span> Casos asignados a {esc(filtro)}</td></tr>"
        if filtro
        else ""
    )

    meta_html = f"""<table class="report-meta" width="100%" cellpadding="0" cellspacing="0">
      <tr>
        <td class="meta-cell" width="50%"><span class="meta-label">Generado</span> {esc(data.get('generadoEn'))}</td>
        <td class="meta-cell" width="50%"><span class="meta-label">Elaborado por</span> {esc(data.get('generadoPor'))}</td>
      </tr>
      {filtro_row}
    </table>"""

    def table_rows(rows: list[dict]) -> str:
        if not rows:
            return "<tr class='empty-row'><td colspan='2'>Sin datos</td></tr>"
        return "".join(
            f"<tr><td>{esc(r.get('label', ''))}</td>"
            f"<td class='num' align='right'>{esc(r.get('total', 0))}</td></tr>"
            for r in rows
        )

    def summary_table(title: str, label_col: str, rows: list[dict]) -> str:
        body = table_rows(rows)
        return f"""
        <div class="section-title section-title--compact">{esc(title)}</div>
        <table class="data summary-table" width="100%" cellpadding="0" cellspacing="0">
          <colgroup><col width="78%" /><col width="22%" /></colgroup>
          <thead><tr><th>{esc(label_col)}</th><th class="th-total" align="right">Total</th></tr></thead>
          <tbody>{body}</tbody>
        </table>"""

    detalle_rows = ""
    for row in data.get("detalle", []):
        semaforo = esc(row.get("semaforo"))
        semaforo_class = ""
        if semaforo == "Vencido":
            semaforo_class = "badge badge--danger"
        elif semaforo == "Próximo a vencer":
            semaforo_class = "badge badge--warn"
        elif semaforo == "Al día":
            semaforo_class = "badge badge--ok"

        semaforo_cell = (
            f"<span class='{semaforo_class}'>{semaforo}</span>" if semaforo_class else semaforo
        )

        detalle_rows += (
            "<tr>"
            f"<td class='mono'>{esc(row.get('radicado'))}</td>"
            f"<td>{esc(row.get('peticionario'))}</td>"
            f"<td>{esc(row.get('estado'))}</td>"
            f"<td>{esc(row.get('recurso'))}</td>"
            f"<td>{esc(row.get('asignado'))}</td>"
            f"<td>{esc(row.get('fechaVencimiento'))}</td>"
            f"<td>{semaforo_cell}</td>"
            "</tr>"
        )

    if not detalle_rows:
        detalle_rows = "<tr class='empty-row'><td colspan='7'>Sin casos</td></tr>"

    def kpi_cell(label: str, value, extra: str = "") -> str:
        return (
            f'<td class="kpi-cell {extra}" width="33%">'
            f'<div class="kpi-label">{esc(label)}</div>'
            f'<div class="kpi-value">{esc(value)}</div>'
            f"</td>"
        )

    kpi_html = f"""<table class="kpi-table" width="100%" cellpadding="0" cellspacing="0">
      <tr>
        {kpi_cell("Total casos", kpis.get("total", 0))}
        {kpi_cell("Pend. radicación", kpis.get("pendienteRadicacion", 0))}
        {kpi_cell("En gestión", kpis.get("enGestion", 0))}
      </tr>
      <tr>
        {kpi_cell("Finalizados", kpis.get("finalizados", 0))}
        {kpi_cell("Vencidos", kpis.get("vencidos", 0), "kpi-danger")}
        {kpi_cell("Próx. a vencer", kpis.get("proximosVencer", 0), "kpi-warn")}
      </tr>
    </table>"""

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8" />
  <title>{esc(data.get('titulo', 'Reporte gerencial'))}</title>
  <style>
    @page {{
      size: A4 portrait;
      margin: 1.4cm 1.6cm 1.6cm 1.6cm;
    }}
    body {{
      font-family: Arial, Helvetica, sans-serif;
      color: #1e293b;
      margin: 0;
      padding: 0;
      font-size: 10pt;
      line-height: 1.35;
      background: #ffffff;
    }}
    .page {{
      width: 100%;
    }}
    .report-header {{
      width: 100%;
      border-bottom: 2px solid #2a5934;
      margin: 0 0 10pt 0;
      padding: 0 0 10pt 0;
    }}
    .report-header__logo {{
      width: 64px;
      vertical-align: middle;
      padding: 0 12pt 0 0;
    }}
    .report-header__text {{
      vertical-align: middle;
      text-align: left;
    }}
    .report-title {{
      margin: 0;
      color: #2a5934;
      font-size: 14pt;
      font-weight: bold;
      line-height: 1.2;
    }}
    .report-subtitle {{
      margin: 3pt 0 0;
      color: #64748b;
      font-size: 9pt;
      line-height: 1.3;
    }}
    .report-meta {{
      width: 100%;
      margin: 0 0 12pt 0;
      border: 1px solid #cbd5e1;
      background: #f8fafc;
      border-collapse: collapse;
    }}
    .meta-cell {{
      padding: 7pt 10pt;
      font-size: 9pt;
      border: 1px solid #e2e8f0;
      vertical-align: top;
    }}
    .meta-label {{
      color: #475569;
      font-weight: bold;
    }}
    .section-title {{
      color: #2a5934;
      font-size: 10.5pt;
      font-weight: bold;
      margin: 12pt 0 6pt 0;
      padding: 0 0 4pt 0;
      border-bottom: 1px solid #cbd5e1;
    }}
    .section-title--compact {{
      font-size: 9.5pt;
      margin: 0 0 4pt 0;
      padding: 0 0 3pt 0;
    }}
    .kpi-table {{
      width: 100%;
      margin: 0 0 10pt 0;
      border-collapse: collapse;
      table-layout: fixed;
    }}
    .kpi-cell {{
      border: 1px solid #cbd5e1;
      background: #ffffff;
      padding: 8pt 10pt;
      vertical-align: top;
    }}
    .kpi-label {{
      font-size: 8pt;
      color: #64748b;
      margin: 0 0 3pt 0;
      line-height: 1.2;
    }}
    .kpi-value {{
      font-size: 15pt;
      font-weight: bold;
      color: #2a5934;
      line-height: 1.1;
      margin: 0;
    }}
    .kpi-danger .kpi-value {{ color: #b91c1c; }}
    .kpi-warn .kpi-value {{ color: #a16207; }}
    .summary-grid {{
      width: 100%;
      border-collapse: collapse;
      margin: 0 0 8pt 0;
    }}
    .summary-grid > tbody > tr > td {{
      width: 50%;
      vertical-align: top;
      padding: 0 8pt 8pt 0;
    }}
    .summary-grid > tbody > tr > td + td {{
      padding: 0 0 8pt 8pt;
    }}
    table.data {{
      width: 100%;
      border-collapse: collapse;
      table-layout: fixed;
      margin: 0;
    }}
    table.data th,
    table.data td {{
      border: 1px solid #cbd5e1;
      padding: 5pt 7pt;
      font-size: 9pt;
      vertical-align: middle;
      word-wrap: break-word;
    }}
    table.data th {{
      background: #2a5934;
      color: #ffffff;
      font-weight: bold;
      text-align: left;
    }}
    table.data td.num,
    table.data th.th-total {{
      text-align: right;
      white-space: nowrap;
    }}
    table.data tbody tr:nth-child(even) td {{
      background: #f8fafc;
    }}
    .detail-table th {{
      font-size: 8pt;
      padding: 5pt 4pt;
    }}
    .detail-table td {{
      font-size: 8pt;
      padding: 4pt 4pt;
      vertical-align: top;
    }}
    .mono {{
      font-family: "Courier New", Courier, monospace;
      font-size: 7.5pt;
    }}
    .badge {{
      display: inline-block;
      padding: 1pt 6pt;
      font-size: 7.5pt;
      font-weight: bold;
      white-space: nowrap;
    }}
    .badge--danger {{ background: #fee2e2; color: #b91c1c; }}
    .badge--warn {{ background: #fef3c7; color: #a16207; }}
    .badge--ok {{ background: #dcfce7; color: #166534; }}
    .empty-row td {{
      text-align: center;
      color: #64748b;
      font-style: italic;
    }}
    .report-footer {{
      margin-top: 14pt;
      padding-top: 6pt;
      border-top: 1px solid #cbd5e1;
      font-size: 7.5pt;
      color: #94a3b8;
      text-align: center;
    }}
  </style>
</head>
<body>
  <div class="page">
  <table class="report-header" width="100%" cellpadding="0" cellspacing="0">
    <tr>
      <td class="report-header__logo" width="72">{logo_html(data)}</td>
      <td class="report-header__text">
        <h1 class="report-title">{esc(data.get('titulo'))}</h1>
        <p class="report-subtitle">{esc(data.get('subtitulo'))}</p>
      </td>
    </tr>
  </table>

  {meta_html}

  <div class="section-title">Indicadores generales</div>
  {kpi_html}

  <table class="summary-grid" width="100%" cellpadding="0" cellspacing="0">
    <tr>
      <td width="50%">{summary_table('Resumen por estado', 'Estado', data.get('porEstado', []))}</td>
      <td width="50%">{summary_table('Resumen por recurso / tema', 'Recurso', data.get('porRecurso', []))}</td>
    </tr>
    <tr>
      <td width="50%">{summary_table('Resumen por canal de reporte', 'Canal', data.get('porCanal', []))}</td>
      <td width="50%">{summary_table('Semáforo de vencimiento', 'Semáforo', data.get('porSemaforo', []))}</td>
    </tr>
  </table>

  <div class="section-title">Detalle de casos</div>
  <table class="data detail-table" width="100%" cellpadding="0" cellspacing="0">
    <colgroup>
      <col width="14%" /><col width="18%" /><col width="11%" /><col width="10%" />
      <col width="16%" /><col width="11%" /><col width="14%" />
    </colgroup>
    <thead>
      <tr>
        <th>Radicado</th><th>Peticionario</th><th>Estado</th><th>Recurso</th>
        <th>Asignado</th><th>Vencimiento</th><th>Semáforo</th>
      </tr>
    </thead>
    <tbody>{detalle_rows}</tbody>
  </table>

  <div class="report-footer">
    Alcaldía de Envigado — Secretaría de Medio Ambiente · CRM ambiental
  </div>
  </div>
</body>
</html>"""


def write_xlsx(data: dict, output_path: Path) -> None:
    if Workbook is None:
        print("Falta openpyxl.", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    header_fill = PatternFill("solid", fgColor="2A5934")
    header_font = Font(color="FFFFFF", bold=True)

    ws["A1"] = data.get("titulo", "Reporte gerencial")
    ws["A1"].font = Font(size=14, bold=True, color="2A5934")
    ws["A2"] = data.get("subtitulo", "")
    ws["A3"] = f"Generado: {data.get('generadoEn', '')}"
    ws["A4"] = f"Elaborado por: {data.get('generadoPor', '')}"

    row = 6
    ws.cell(row=row, column=1, value="Indicador").fill = header_fill
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=2, value="Total").fill = header_fill
    ws.cell(row=row, column=2).font = header_font

    kpis = data.get("kpis", {})
    indicadores = [
        ("Total casos", kpis.get("total", 0)),
        ("Pendientes de radicación", kpis.get("pendienteRadicacion", 0)),
        ("En gestión", kpis.get("enGestion", 0)),
        ("Finalizados", kpis.get("finalizados", 0)),
        ("Vencidos (activos)", kpis.get("vencidos", 0)),
        ("Próximos a vencer", kpis.get("proximosVencer", 0)),
    ]

    for label, total in indicadores:
        row += 1
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=total)

    def write_sheet(name: str, rows: list[dict]) -> None:
        sheet = wb.create_sheet(name[:31])
        sheet.append([name.split(" ", 1)[0], "Total"])
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for item in rows:
            sheet.append([item.get("label", ""), item.get("total", 0)])

    write_sheet("Por estado", data.get("porEstado", []))
    write_sheet("Por recurso", data.get("porRecurso", []))
    write_sheet("Por canal", data.get("porCanal", []))
    write_sheet("Por semaforo", data.get("porSemaforo", []))

    detalle = wb.create_sheet("Detalle casos")
    detalle.append([
        "Radicado", "Expediente", "Peticionario", "Estado", "Recurso", "Canal",
        "Asignado", "Fecha caso", "Vencimiento", "Semáforo", "Barrio",
    ])
    for cell in detalle[1]:
        cell.fill = header_fill
        cell.font = header_font

    for item in data.get("detalle", []):
        detalle.append([
            item.get("radicado"),
            item.get("expediente"),
            item.get("peticionario"),
            item.get("estado"),
            item.get("recurso"),
            item.get("canal"),
            item.get("asignado"),
            item.get("fechaCaso"),
            item.get("fechaVencimiento"),
            item.get("semaforo"),
            item.get("barrio"),
        ])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[letter].width = min(max_len + 2, 42)

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--payload", required=True)
    parser.add_argument("--format", required=True, choices=["pdf", "xlsx"])
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    payload_path = Path(args.payload)
    output_base = Path(args.output)

    with payload_path.open("r", encoding="utf-8") as fh:
        data = json.load(fh)

    if args.format == "xlsx":
        write_xlsx(data, output_base.with_suffix(".xlsx"))
        return

    html_content = build_html(data)
    output_base.with_suffix(".html").write_text(html_content, encoding="utf-8")


if __name__ == "__main__":
    main()
