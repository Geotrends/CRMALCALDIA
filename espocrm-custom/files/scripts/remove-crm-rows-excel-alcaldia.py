#!/usr/bin/env python3
"""
Elimina filas del Excel oficial que fueron escritas por el CRM.
Conserva el histórico que ya venía en excelAlcaldia.xlsx.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Falta openpyxl.", file=sys.stderr)
    sys.exit(1)

SHEET_NAME = "2026"
HEADER_ROW = 1
DATA_START_ROW = 2


def normalize_header(value) -> str:
    return str(value or "").strip().lower()


def build_header_index(ws) -> dict[str, list[int]]:
    index: dict[str, list[int]] = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=HEADER_ROW, column=col).value
        if header is None:
            continue
        index.setdefault(normalize_header(header), []).append(col)
    return index


def find_column(headers: dict[str, list[int]], header_name: str, occurrence: int = 1) -> int | None:
    key = normalize_header(header_name)
    cols = headers.get(key, [])
    if occurrence <= len(cols):
        return cols[occurrence - 1]
    return None


def remove_rows(excel_path: Path, radicados: set[str], expedientes: set[str]) -> int:
    wb = load_workbook(excel_path)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"No existe la hoja {SHEET_NAME}")

    ws = wb[SHEET_NAME]
    headers = build_header_index(ws)
    rad_col = find_column(headers, "Radicado")
    con_col = find_column(headers, "Consecutivo")

    if not rad_col and not con_col:
        raise ValueError("No se encontraron columnas Radicado/Consecutivo")

    rows_to_delete: list[int] = []

    for row in range(DATA_START_ROW, ws.max_row + 1):
        rad = str(ws.cell(row=row, column=rad_col).value or "").strip() if rad_col else ""
        con = str(ws.cell(row=row, column=con_col).value or "").strip() if con_col else ""

        if rad in radicados or rad.startswith("ENV-"):
            rows_to_delete.append(row)
            continue

        if con in expedientes:
            rows_to_delete.append(row)

    for row in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row, 1)

    wb.save(excel_path)

    return len(rows_to_delete)


def main() -> None:
    if len(sys.argv) < 2:
        print("Uso: remove-crm-rows-excel-alcaldia.py <excelAlcaldia.xlsx>", file=sys.stderr)
        sys.exit(1)

    payload = json.load(sys.stdin)
    radicados = {str(v).strip() for v in payload.get("radicados", []) if str(v).strip()}
    expedientes = {str(v).strip() for v in payload.get("expedientes", []) if str(v).strip()}

    removed = remove_rows(Path(sys.argv[1]), radicados, expedientes)
    print(f"Filas CRM eliminadas: {removed}")


if __name__ == "__main__":
    main()
