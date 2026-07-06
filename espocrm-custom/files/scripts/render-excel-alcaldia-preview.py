#!/usr/bin/env python3
"""
Genera HTML de vista previa para excelAlcaldia.xlsx (hoja 2026).
Salida: JSON en stdout con sheetName y html.
"""

from __future__ import annotations

import html
import json
import sys
from datetime import date, datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "missing_openpyxl"}), file=sys.stderr)
    sys.exit(1)

SHEET_NAME = "2026"
MAX_ROWS = 2500


def format_cell(value) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    return str(value)


def build_html(rows: list[tuple]) -> str:
    if not rows:
        return ""

    max_cols = max(len(row) for row in rows)
    parts = ['<table id="excel-alcaldia-sheet-table"><thead><tr>']
    header = rows[0]

    for index in range(max_cols):
        value = header[index] if index < len(header) else ""
        parts.append("<th>" + html.escape(format_cell(value)) + "</th>")

    parts.append("</tr></thead><tbody>")

    for row in rows[1:]:
        parts.append("<tr>")

        for index in range(max_cols):
            value = row[index] if index < len(row) else ""
            parts.append("<td>" + html.escape(format_cell(value)) + "</td>")

        parts.append("</tr>")

    parts.append("</tbody></table>")

    return "".join(parts)


def main() -> int:
    if len(sys.argv) < 2:
        print("Uso: render-excel-alcaldia-preview.py <excelAlcaldia.xlsx>", file=sys.stderr)
        return 1

    excel_path = Path(sys.argv[1])

    if not excel_path.is_file():
        print(json.dumps({"error": "file_not_found"}))
        return 1

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet_name = SHEET_NAME if SHEET_NAME in workbook.sheetnames else workbook.sheetnames[0]
    worksheet = workbook[sheet_name]

    rows: list[tuple] = []

    for index, row in enumerate(worksheet.iter_rows(values_only=True)):
        if index >= MAX_ROWS:
            break

        rows.append(tuple(row))

    while rows and not any(
        cell is not None and str(cell).strip() != "" for cell in rows[-1]
    ):
        rows.pop()

    workbook.close()

    payload = {
        "sheetName": sheet_name,
        "html": build_html(rows),
        "rowCount": max(0, len(rows) - 1),
    }

    print(json.dumps(payload, ensure_ascii=False))

    return 0


if __name__ == "__main__":
    sys.exit(main())
