#!/usr/bin/env python3
"""
Excel maestro único: una fila por caso.
- section=solicitud → columnas del Case (al radicar)
- section=acta     → columnas del ActaVisita (misma fila, no borra solicitud)
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("Falta openpyxl.", file=sys.stderr)
    sys.exit(1)

ID_FIELD = "id"

SOLICITUD_LABELS = {
    "cFechaCaso": "Fecha del caso",
    "cNumeroRadicado": "Número de radicado",
    "cExpediente": "Expediente / consecutivo interno",
    "cPeticionario": "Peticionario",
    "cCedula": "Cédula del peticionario",
    "cDireccion": "Dirección del peticionario",
    "cTelefono": "Teléfono del peticionario",
    "cBarrio": "Barrio del peticionario",
    "cCorreo": "Correo electrónico",
    "cCanalDeReporte": "Canal de reporte",
    "cTipo": "Tipo de solicitud",
    "cCategoria": "Categoría ambiental",
    "cPerjudicante": "Perjudicante / posible afectante",
    "cTelefonoPerjudicante": "Teléfono del perjudicante",
    "cDireccionPerjudicante": "Dirección del perjudicante",
    "cBarrioPerjudicante": "Barrio del perjudicante",
    "description": "Descripción de la queja / reporte",
    "cRespuestaInmediata": "Respuesta inmediata",
    "cRecibidaPor": "Recibida por",
    "cRemitidoA": "Remitido a",
    "assignedUser": "Asignado a (patrullero)",
}

ACTA_LABELS = {
    "fechaVisita": "Fecha de la visita",
    "anio": "Año acta",
    "autorizacionDatos": "Autorización datos personales",
    "posibleAfectante": "Posible afectante",
    "direccionAfectacion": "Dirección afectación",
    "telefono": "Teléfono (acta)",
    "barrio": "Barrio (acta)",
    "zona": "Zona",
    "objetoVisita": "Objeto de la visita",
    "situacionEncontrada": "Situación encontrada",
    "analisisSituacion": "Análisis de la situación",
    "conclusion": "Conclusión",
    "requerimientos": "Requerimientos",
    "funcionarioNombre": "Nombre funcionario",
    "funcionarioCedula": "C.C. funcionario",
    "funcionarioCargo": "Cargo funcionario",
    "establecimientoNombre": "Nombre establecimiento",
    "establecimientoCedula": "C.C. establecimiento",
    "establecimientoCargo": "Cargo establecimiento",
}

ALL_COLUMNS: list[tuple[str, str, str]] = (
    [(ID_FIELD, "ID caso", "solicitud")]
    + [(c, SOLICITUD_LABELS[c], "solicitud") for c in SOLICITUD_LABELS]
    + [(c, ACTA_LABELS[c], "acta") for c in ACTA_LABELS]
)

FILL_SOL = PatternFill("solid", fgColor="1F4E79")
FILL_ACTA = PatternFill("solid", fgColor="2E7D32")
FONT_HEADER = Font(bold=True, color="FFFFFF")
FONT_CODE = Font(italic=True, color="444444")


def column_map(ws) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        code = ws.cell(row=2, column=col).value
        if code:
            mapping[str(code).strip()] = col
    return mapping


def ensure_structure(ws) -> None:
    existing = column_map(ws)
    expected_codes = [c[0] for c in ALL_COLUMNS]

    if not existing:
        for col, (code, label, section) in enumerate(ALL_COLUMNS, start=1):
            fill = FILL_SOL if section == "solicitud" else FILL_ACTA
            h = ws.cell(row=1, column=col, value=label)
            h.fill = fill
            h.font = FONT_HEADER
            h.alignment = Alignment(horizontal="center", wrap_text=True)
            c = ws.cell(row=2, column=col, value=code)
            c.font = FONT_CODE
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[h.column_letter].width = max(14, min(28, len(label) * 0.85))
        ws.title = "Registro CRM"
        ws.row_dimensions[1].height = 42
        ws.freeze_panes = "A3"
        return

    next_col = ws.max_column + 1
    for code, label, section in ALL_COLUMNS:
        if code in existing:
            continue
        fill = FILL_SOL if section == "solicitud" else FILL_ACTA
        h = ws.cell(row=1, column=next_col, value=label)
        h.fill = fill
        h.font = FONT_HEADER
        ws.cell(row=2, column=next_col, value=code).font = FONT_CODE
        ws.column_dimensions[h.column_letter].width = max(14, min(28, len(label) * 0.85))
        existing[code] = next_col
        next_col += 1


def find_row_by_case_id(ws, case_id: str) -> int | None:
    cols = column_map(ws)
    id_col = cols.get(ID_FIELD, 1)
    for row in range(3, ws.max_row + 1):
        value = ws.cell(row=row, column=id_col).value
        if value is not None and str(value).strip() == case_id:
            return row
    return None


def next_data_row(ws) -> int:
    cols = column_map(ws)
    id_col = cols.get(ID_FIELD, 1)
    row = 3
    while row <= ws.max_row:
        if ws.cell(row=row, column=id_col).value in (None, ""):
            return row
        row += 1
    return ws.max_row + 1 if ws.max_row >= 2 else 3


def upsert(excel_path: Path, payload: dict) -> None:
    case_id = str(payload.get("id") or "").strip()
    if not case_id:
        raise ValueError("Falta id del caso.")

    fields = payload.get("fields") or {}
    if not fields:
        return

    if excel_path.is_file():
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        excel_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active

    ensure_structure(ws)
    cols = column_map(ws)

    row = find_row_by_case_id(ws, case_id)
    if row is None:
        row = next_data_row(ws)
        if ID_FIELD in cols:
            ws.cell(row=row, column=cols[ID_FIELD], value=case_id)

    for code, value in fields.items():
        if code == ID_FIELD:
            continue
        col = cols.get(code)
        if not col:
            continue
        ws.cell(row=row, column=col, value="" if value is None else value)

    wb.save(excel_path)


def main() -> None:
    if len(sys.argv) < 2:
        print("Uso: upsert-crm-excel.py <ruta.xlsx>", file=sys.stderr)
        sys.exit(1)

    upsert(Path(sys.argv[1]), json.load(sys.stdin))
    print(sys.argv[1])


if __name__ == "__main__":
    main()
