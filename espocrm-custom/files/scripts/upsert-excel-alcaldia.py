#!/usr/bin/env python3
"""
Escribe/actualiza una fila en excelAlcaldia.xlsx (hoja 2026).

Casos del CRM (radicado ENV-…):
  - Se ubican en bloque al final del histórico (justo debajo de las filas existentes).
  - Cada caso nuevo se apila consecutivamente en ese bloque.
  - Si ya existe el radicado, se actualiza la misma fila.
"""

from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Falta openpyxl.", file=sys.stderr)
    sys.exit(1)

SHEET_NAME = "2026"
HEADER_ROW = 1
DATA_START_ROW = 2
CRM_RADICADO_PREFIX = "ENV-"
# Filas vacías seguidas → fin del bloque histórico contiguo (evita saltar a fila 900+).
HISTORICAL_GAP_THRESHOLD = 3

# clave payload -> (encabezado, ocurrencia).
# El Excel repite encabezados (quejoso / infractor); occurrence indica la columna.
PAYLOAD_TO_HEADER = {
    "consecutivo": ("Consecutivo", 1),
    "radicado": ("Radicado", 1),
    "solicitante": ("Solicitante", 1),
    "via_principal_quejoso": ("Vía principal", 1),
    "num_via_principal_quejoso": ("N° vía principal", 1),
    "letra_via_quejoso": ("letra", 1),
    "cuadrante_via_quejoso": ("Cuadrante", 1),
    "generadora_quejoso": ("Generadora", 1),
    "letra_generadora_quejoso": ("Letra", 1),
    "cuadrante_generadora_quejoso": ("Cuadrante", 2),
    "placa_quejoso": ("Placa", 1),
    "bloque_quejoso": ("Bloque", 1),
    "interior_quejoso": ("Interior", 1),
    "direccion_quejoso": ("Dirección", 1),
    "cedula_quejoso": ("Cedula quejoso", 1),
    "telefono_quejoso": ("Teléfono quejoso", 1),
    "correo_quejoso": ("Correo electrónico", 1),
    "infractor": ("Infractor", 1),
    "via_principal_infractor": ("Vía principal", 2),
    "num_via_principal_infractor": ("N° vía principal", 2),
    "letra_via_infractor": ("letra", 3),
    "cuadrante_via_infractor": ("Cuadrante", 3),
    "generadora_infractor": ("Generadora", 2),
    "letra_generadora_infractor": ("Letra", 2),
    "cuadrante_generadora_infractor": ("Cuadrante", 4),
    "placa_infractor": ("Placa", 2),
    "bloque_infractor": ("Bloque", 2),
    "interior_infractor": ("Interior", 2),
    "direccion_infractor": ("Direccion", 1),
    "cedula_infractor": ("Cedula Infractor", 1),
    "telefono_infractor": ("Teléfono Infractor", 1),
    "correo_infractor": ("Correo electrónico", 2),
    "recurso_tema": ("Recurso - Tema", 1),
    "asunto": (" Asunto", 1),
    "barrio": ("Barrio", 1),
    "zona": ("Zona", 1),
    "fecha_ingreso": ("Fecha de ingreso (dd/mm/aaaa)", 1),
    "fecha_vencimiento": ("fecha de ultima actuación", 1),
    "ultima_actuacion": ("Ultima actuación", 1),
    "inspector": ("Inspector responsable", 1),
    "proxima_actuacion": ("Próxima actuación", 1),
    "descripcion": ("Observaciones", 1),
    "canal_reporte": ("Observaciones", 1),
}


def normalize_header(value) -> str:
    return str(value or "").strip().lower()


def build_header_index(ws) -> dict[str, list[int]]:
    index: dict[str, list[int]] = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=HEADER_ROW, column=col).value
        if header is None:
            continue
        key = normalize_header(header)
        index.setdefault(key, []).append(col)
    return index


def find_column(
    headers: dict[str, list[int]],
    header_name: str,
    occurrence: int = 1,
) -> int | None:
    key = normalize_header(header_name)

    if key in headers:
        cols = headers[key]
        if occurrence <= len(cols):
            return cols[occurrence - 1]
        return None

    fuzzy_matches: list[int] = []
    for stored, cols in headers.items():
        if key in stored or stored in key:
            fuzzy_matches.extend(cols)

    if not fuzzy_matches:
        return None

    unique_cols = sorted(set(fuzzy_matches))
    if occurrence <= len(unique_cols):
        return unique_cols[occurrence - 1]

    return None


def find_row(ws, headers: dict[str, int], radicado: str, consecutivo: str) -> int | None:
    rad_col = find_column(headers, "Radicado")
    con_col = find_column(headers, "Consecutivo")

    if radicado and rad_col:
        for row in range(DATA_START_ROW, ws.max_row + 1):
            rad_val = str(ws.cell(row=row, column=rad_col).value or "").strip()
            if rad_val == radicado:
                return row

    if consecutivo and con_col and not radicado:
        for row in range(DATA_START_ROW, ws.max_row + 1):
            con_val = str(ws.cell(row=row, column=con_col).value or "").strip()
            if con_val == consecutivo:
                return row

    return None


def is_crm_radicado(radicado: str) -> bool:
    return str(radicado or "").strip().startswith(CRM_RADICADO_PREFIX)


def row_has_data(ws, headers: dict[str, list[int]], row: int) -> bool:
    for name in ("Radicado", "Consecutivo", "Solicitante"):
        col = find_column(headers, name)
        if not col:
            continue
        value = ws.cell(row=row, column=col).value
        if value is not None and str(value).strip() != "":
            return True
    return False


def last_historical_row(ws, headers: dict[str, list[int]]) -> int:
    """Última fila con datos contiguos al inicio (sin saltar huecos vacíos al final)."""
    rad_col = find_column(headers, "Radicado")
    last = DATA_START_ROW - 1
    empty_streak = 0

    for row in range(DATA_START_ROW, ws.max_row + 1):
        if rad_col:
            rad = str(ws.cell(row=row, column=rad_col).value or "").strip()
            if rad.startswith(CRM_RADICADO_PREFIX):
                break

        if row_has_data(ws, headers, row):
            last = row
            empty_streak = 0
            continue

        if last < DATA_START_ROW:
            continue

        empty_streak += 1
        if empty_streak >= HISTORICAL_GAP_THRESHOLD:
            break

    return last


def last_crm_row(ws, headers: dict[str, list[int]]) -> int | None:
    rad_col = find_column(headers, "Radicado")
    if not rad_col:
        return None

    last = None
    for row in range(DATA_START_ROW, ws.max_row + 1):
        rad = str(ws.cell(row=row, column=rad_col).value or "").strip()
        if rad.startswith(CRM_RADICADO_PREFIX):
            last = row

    return last


def next_crm_row(ws, headers: dict[str, list[int]]) -> int:
    """Primera fila libre del bloque CRM: debajo del histórico o del último ENV-."""
    crm_last = last_crm_row(ws, headers)
    if crm_last is not None:
        return crm_last + 1
    return last_historical_row(ws, headers) + 1


def last_data_row(ws, headers: dict[str, list[int]]) -> int:
    key_cols: list[int] = []
    for name in ("Radicado", "Consecutivo", "Solicitante"):
        col = find_column(headers, name)
        if col:
            key_cols.append(col)

    if not key_cols:
        return DATA_START_ROW - 1

    last = DATA_START_ROW - 1
    for row in range(DATA_START_ROW, ws.max_row + 1):
        for col in key_cols:
            value = ws.cell(row=row, column=col).value
            if value is not None and str(value).strip() != "":
                last = row
                break

    return last


def next_empty_row(ws, headers: dict[str, list[int]]) -> int:
    return last_data_row(ws, headers) + 1


def merge_observaciones(existing: str, parts: list[str]) -> str:
    chunks = [existing.strip()] if existing and str(existing).strip() else []
    for part in parts:
        part = str(part or "").strip()
        if part and part not in chunks:
            chunks.append(part)
    return " | ".join(chunks)


def upsert(excel_path: Path, payload: dict) -> None:
    radicado = str(payload.get("radicado") or "").strip()
    consecutivo = str(payload.get("consecutivo") or "").strip()

    if not radicado and not consecutivo:
        raise ValueError("Falta radicado o consecutivo para ubicar la fila.")

    if not excel_path.is_file():
        raise FileNotFoundError(f"No existe el Excel oficial: {excel_path}")

    wb = load_workbook(excel_path)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"No existe la hoja {SHEET_NAME} en {excel_path}")

    ws = wb[SHEET_NAME]
    headers = build_header_index(ws)

    row = find_row(ws, headers, radicado, consecutivo)
    if row is None:
        if is_crm_radicado(radicado):
            row = next_crm_row(ws, headers)
        else:
            row = next_empty_row(ws, headers)

    observaciones_parts = []
    if payload.get("descripcion"):
        observaciones_parts.append("Queja: " + str(payload["descripcion"]).strip())
    if payload.get("canal_reporte"):
        observaciones_parts.append("Canal: " + str(payload["canal_reporte"]).strip())

    for key, spec in PAYLOAD_TO_HEADER.items():
        if key in ("descripcion", "canal_reporte"):
            continue
        header, occurrence = spec
        col = find_column(headers, header, occurrence)
        if not col:
            continue
        if key not in payload:
            continue
        value = payload.get(key)
        if value is None:
            continue
        ws.cell(row=row, column=col, value=str(value).strip() if value != "" else "")

    obs_col = find_column(headers, "Observaciones")
    if obs_col and observaciones_parts:
        existing = ws.cell(row=row, column=obs_col).value
        ws.cell(row=row, column=obs_col, value=merge_observaciones(existing, observaciones_parts))

    mod_col = find_column(headers, "Fecha ultima modificación")
    if mod_col:
        ws.cell(row=row, column=mod_col, value=datetime.now().strftime("%d/%m/%Y %H:%M"))

    wb.save(excel_path)


def main() -> None:
    if len(sys.argv) < 2:
        print("Uso: upsert-excel-alcaldia.py <excelAlcaldia.xlsx>", file=sys.stderr)
        sys.exit(1)

    upsert(Path(sys.argv[1]), json.load(sys.stdin))
    print(sys.argv[1])


if __name__ == "__main__":
    main()
