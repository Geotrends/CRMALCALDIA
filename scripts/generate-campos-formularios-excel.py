#!/usr/bin/env python3
"""
Genera un Excel con los campos de los formularios:
  - Formato solicitud (Case)
  - Acta de visita (ActaVisita)

Uso:
  python3 scripts/generate-campos-formularios-excel.py
  python3 scripts/generate-campos-formularios-excel.py -o exports/mi-archivo.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Falta openpyxl. Instala con: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
CUSTOM = ROOT / "espocrm-custom" / "Resources"
OUTPUT_DEFAULT = ROOT / "exports" / "campos-formularios-crm.xlsx"

# Etiquetas del Case (no todas están en i18n custom)
CASE_FIELD_LABELS = {
    "cFechaCaso": "Fecha del caso",
    "cNumeroRadicado": "Número de radicado",
    "cExpediente": "Expediente / consecutivo interno",
    "cPeticionario": "Peticionario",
    "cCedula": "Cédula del peticionario",
    "cDireccion": "Dirección del peticionario",
    "cTelefono": "Teléfono del peticionario",
    "cBarrio": "Barrio del peticionario",
    "cCorreo": "Correo electrónico",
    "cCanalDeReporte": "Canal de reporte",
    "cTipo": "Tipo de solicitud",
    "cCategoria": "Categoría ambiental",
    "cPerjudicante": "Perjudicante / posible afectante",
    "cTelefonoPerjudicante": "Teléfono del perjudicante",
    "cDireccionPerjudicante": "Dirección del perjudicante",
    "cBarrioPerjudicante": "Barrio del perjudicante",
    "description": "Descripción de la queja / reporte",
    "cRespuestaInmediata": "Respuesta inmediata",
    "cRecibidaPor": "Recibida por",
    "cRemitidoA": "Remitido a",
    "assignedUser": "Asignado a (patrullero)",
    "status": "Estado del caso",
    "name": "Nombre / asunto del caso",
}

CASE_FIELD_TYPES = {
    "cFechaCaso": "fecha y hora",
    "cNumeroRadicado": "texto",
    "cExpediente": "texto",
    "cPeticionario": "texto",
    "cCedula": "texto",
    "cDireccion": "texto",
    "cTelefono": "texto",
    "cBarrio": "texto",
    "cCorreo": "correo",
    "cCanalDeReporte": "lista",
    "cTipo": "lista",
    "cCategoria": "lista múltiple",
    "cPerjudicante": "texto",
    "cTelefonoPerjudicante": "texto",
    "cDireccionPerjudicante": "texto",
    "cBarrioPerjudicante": "texto",
    "description": "texto largo",
    "cRespuestaInmediata": "texto largo",
    "cRecibidaPor": "usuario",
    "cRemitidoA": "usuario",
    "assignedUser": "usuario",
    "status": "lista (estado)",
    "name": "texto",
}


def load_json(path: Path) -> dict | list:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "utf-16", "utf-16-le"):
        try:
            return json.loads(raw.decode(encoding))
        except (UnicodeDecodeError, json.JSONDecodeError):
            continue
    raise ValueError(f"No se pudo leer JSON: {path}")


def extract_fields_from_layout(layout: list) -> list[tuple[str, str]]:
    """Devuelve [(seccion, nombre_campo), ...] en orden del formulario."""
    result: list[tuple[str, str]] = []
    for panel in layout:
        if not isinstance(panel, dict):
            continue
        section = panel.get("label") or panel.get("name") or "General"
        for row in panel.get("rows") or []:
            if not isinstance(row, list):
                continue
            for cell in row:
                if not isinstance(cell, dict):
                    continue
                name = cell.get("name")
                if name:
                    result.append((section, name))
    return result


def load_acta_labels() -> dict[str, str]:
    i18n_path = CUSTOM / "i18n" / "es_ES" / "ActaVisita.json"
    if i18n_path.is_file():
        data = load_json(i18n_path)
        return data.get("fields") or {}
    return {}


def load_acta_types() -> dict[str, str]:
    defs_path = CUSTOM / "metadata" / "entityDefs" / "ActaVisita.json"
    if not defs_path.is_file():
        return {}
    fields = load_json(defs_path).get("fields") or {}
    type_map = {
        "varchar": "texto",
        "text": "texto largo",
        "int": "número",
        "bool": "sí/no",
        "date": "fecha",
        "datetime": "fecha y hora",
        "enum": "lista",
        "link": "relación",
        "file": "archivo",
        "attachmentMultiple": "archivos / fotos",
    }
    result = {}
    for name, meta in fields.items():
        raw = meta.get("type", "")
        result[name] = type_map.get(raw, raw)
    return result


def build_rows_solicitud() -> list[dict]:
    layout_path = CUSTOM / "layouts" / "Case" / "edit.json"
    layout = load_json(layout_path)
    fields = extract_fields_from_layout(layout)

    rows = []
    for i, (section, name) in enumerate(fields, start=1):
        rows.append(
            {
                "orden": i,
                "seccion": section,
                "codigo_campo": name,
                "etiqueta": CASE_FIELD_LABELS.get(name, name),
                "tipo": CASE_FIELD_TYPES.get(name, ""),
                "entidad": "Case (Caso)",
                "formulario": "Formato solicitud",
            }
        )
    return rows


def build_rows_acta() -> list[dict]:
    layout_path = CUSTOM / "layouts" / "ActaVisita" / "edit.json"
    layout = load_json(layout_path)
    fields = extract_fields_from_layout(layout)
    labels = load_acta_labels()
    types = load_acta_types()

    rows = []
    for i, (section, name) in enumerate(fields, start=1):
        rows.append(
            {
                "orden": i,
                "seccion": section,
                "codigo_campo": name,
                "etiqueta": labels.get(name, name),
                "tipo": types.get(name, ""),
                "entidad": "ActaVisita",
                "formulario": "Acta de visita",
            }
        )
    return rows


def write_horizontal_sheet(ws, solicitud: list[dict], acta: list[dict]) -> None:
    """Una fila de encabezados: primero solicitud, luego acta (horizontal)."""
    ws.title = "Campos formularios"

    fill_solicitud = PatternFill("solid", fgColor="1F4E79")
    fill_acta = PatternFill("solid", fgColor="2E7D32")
    font_header = Font(bold=True, color="FFFFFF")
    font_code = Font(italic=True, color="444444")

    col = 1

    # Fila 1: etiquetas
    for item in solicitud:
        cell = ws.cell(row=1, column=col, value=item["etiqueta"])
        cell.fill = fill_solicitud
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = max(14, min(28, len(item["etiqueta"]) * 0.9))
        col += 1

    for item in acta:
        cell = ws.cell(row=1, column=col, value=item["etiqueta"])
        cell.fill = fill_acta
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = max(14, min(28, len(item["etiqueta"]) * 0.9))
        col += 1

    # Fila 2: código interno (referencia)
    col = 1
    for item in solicitud:
        cell = ws.cell(row=2, column=col, value=item["codigo_campo"])
        cell.font = font_code
        cell.alignment = Alignment(horizontal="center")
        col += 1

    for item in acta:
        cell = ws.cell(row=2, column=col, value=item["codigo_campo"])
        cell.font = font_code
        cell.alignment = Alignment(horizontal="center")
        col += 1

    ws.row_dimensions[1].height = 48
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "A3"


def write_horizontal_sheet_single(ws, title: str, rows: list[dict], fill_color: str) -> None:
    """Hoja con una sola fila horizontal de campos."""
    ws.title = title[:31]
    fill = PatternFill("solid", fgColor=fill_color)
    font_header = Font(bold=True, color="FFFFFF")
    font_code = Font(italic=True, color="444444")

    for col, item in enumerate(rows, start=1):
        cell = ws.cell(row=1, column=col, value=item["etiqueta"])
        cell.fill = fill
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = max(14, min(28, len(item["etiqueta"]) * 0.9))

        code_cell = ws.cell(row=2, column=col, value=item["codigo_campo"])
        code_cell.font = font_code
        code_cell.alignment = Alignment(horizontal="center")

    ws.row_dimensions[1].height = 48
    ws.freeze_panes = "A3"


def main() -> None:
    parser = argparse.ArgumentParser(description="Genera Excel con campos de formularios CRM")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=OUTPUT_DEFAULT,
        help=f"Ruta del Excel (default: {OUTPUT_DEFAULT})",
    )
    args = parser.parse_args()

    solicitud = build_rows_solicitud()
    acta = build_rows_acta()

    args.output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    write_horizontal_sheet(ws, solicitud, acta)

    ws_sol = wb.create_sheet("Solo solicitud")
    write_horizontal_sheet_single(ws_sol, "Solo solicitud", solicitud, "1F4E79")

    ws_acta = wb.create_sheet("Solo acta visita")
    write_horizontal_sheet_single(ws_acta, "Solo acta visita", acta, "2E7D32")

    wb.save(args.output)

    print(f"Excel creado: {args.output}")
    print(f"  - Hoja principal: 1 fila horizontal (solicitud {len(solicitud)} + acta {len(acta)} campos)")
    print(f"  - Fila 1: etiquetas | Fila 2: código interno | Fila 3+: para diligenciar")


if __name__ == "__main__":
    main()
